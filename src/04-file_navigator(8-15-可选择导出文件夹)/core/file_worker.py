import os
import shutil
import tempfile
import time
import zipfile
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Alignment


class FileWorker:
    """文件处理核心引擎：遍历、导出、HTML生成、空文件夹检测、打包发布"""

    def __init__(self):
        self.file_data = []
        self.empty_folders = []
        self.root_path = None

    # ==================== 快速计数 ====================
    def count_files_fast(self, root_dir, limit=200):
        try:
            count = 0
            root_path = Path(root_dir).resolve()
            for item in root_path.rglob('*'):
                if item.is_file():
                    count += 1
                    if count > limit:
                        return count
            return count
        except (PermissionError, OSError):
            return -1

    # ==================== 扫描与遍历 ====================
    def scan_directory(self, root_dir, extensions=None, detect_empty=True):
        start_time = time.time()
        self.root_path = Path(root_dir).resolve()
        self.file_data = []
        self.empty_folders = []

        ext_set = None
        if extensions:
            ext_set = set()
            for ext in extensions:
                ext = ext.strip()
                if ext and not ext.startswith('.'):
                    ext = '.' + ext
                ext_set.add(ext.lower())

        all_folders = set()
        folders_with_files = set()

        for file_path in self.root_path.rglob('*'):
            try:
                if file_path.is_file():
                    parent_folder = str(file_path.parent)
                    folders_with_files.add(parent_folder)

                    ext = file_path.suffix.lower()
                    if ext_set and ext not in ext_set:
                        continue

                    stat = file_path.stat()
                    size_bytes = stat.st_size

                    if size_bytes < 1024:
                        size_str = f"{size_bytes} B"
                    elif size_bytes < 1024 * 1024:
                        size_str = f"{size_bytes / 1024:.1f} KB"
                    elif size_bytes < 1024 * 1024 * 1024:
                        size_str = f"{size_bytes / (1024 * 1024):.1f} MB"
                    else:
                        size_str = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"

                    mtime = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    rel_path = str(file_path.relative_to(self.root_path))
                    abs_path = str(file_path)

                    self.file_data.append({
                        '文件名': file_path.name,
                        '后缀': ext if ext else '(无后缀)',
                        '完整路径': abs_path,
                        '相对路径': rel_path,
                        '文件大小': size_str,
                        '大小(字节)': size_bytes,
                        '修改时间': mtime,
                        '所在文件夹': str(file_path.parent)
                    })

                elif file_path.is_dir():
                    all_folders.add(str(file_path))

            except (PermissionError, OSError):
                continue

        if detect_empty:
            all_folders.add(str(self.root_path))
            empty_set = all_folders - folders_with_files
            self.empty_folders = sorted([f for f in empty_set if f != str(self.root_path)])

        elapsed_time = time.time() - start_time

        return {
            'files': self.file_data,
            'empty_folders': self.empty_folders,
            'total_count': len(self.file_data),
            'scan_time': elapsed_time
        }

    # ==================== 导出 Excel（支持自定义字段和路径过滤） ====================
    def export_to_excel(self, output_path, selected_fields=None, with_hyperlink=True, include_empty_folders=True, include_paths=None):
        if not self.file_data:
            return False

        all_fields = ['文件名', '后缀', '完整路径', '文件大小', '修改时间', '所在文件夹']
        if not selected_fields:
            selected_fields = all_fields

        # 如果有指定要导出的路径列表，过滤数据
        if include_paths:
            file_data_to_export = [f for f in self.file_data if f['完整路径'] in include_paths]
        else:
            file_data_to_export = self.file_data

        if not file_data_to_export:
            return False

        export_data = []
        for item in file_data_to_export:
            row = {}
            for field in selected_fields:
                if field in item:
                    row[field] = item[field]
            export_data.append(row)

        df = pd.DataFrame(export_data)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='文件清单', index=False)

            if with_hyperlink and '完整路径' in selected_fields:
                workbook = writer.book
                worksheet = workbook['文件清单']

                header_row = 1
                col_idx = None
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=header_row, column=col).value
                    if cell_value == '完整路径':
                        col_idx = col
                        break

                if col_idx:
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        file_path = cell.value
                        if file_path and os.path.exists(file_path):
                            cell.value = f'=HYPERLINK("{file_path}", "{os.path.basename(file_path)}")'
                            cell.font = Font(color='0000FF', underline='single')
                            cell.alignment = Alignment(horizontal='left', vertical='center')

            if include_empty_folders and self.empty_folders:
                empty_df = pd.DataFrame({
                    '空文件夹路径': self.empty_folders
                })
                empty_df.to_excel(writer, sheet_name='空文件夹列表', index=False)

        return True

    # ==================== 导出 HTML ====================
    def export_to_html(self, output_path, title="文件目录树", use_relative_path=False, root_for_relative=None):
        if not self.file_data:
            return False

        tree = self._build_tree(self.file_data)
        html_content = self._render_html_tree(
            tree, title, self.root_path,
            use_relative_path=use_relative_path,
            root_for_relative=root_for_relative or self.root_path
        )

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            return True
        except Exception:
            return False

    def _build_tree(self, files):
        tree = {}
        for item in files:
            rel_path = item.get('相对路径', '')
            if not rel_path:
                continue

            parts = Path(rel_path).parts
            current = tree

            for i, part in enumerate(parts[:-1]):
                if part not in current:
                    current[part] = {'_files': []}
                current = current[part]

            file_name = parts[-1]
            if '_files' not in current:
                current['_files'] = []
            current['_files'].append({
                'name': file_name,
                'path': item['完整路径'],
                'size': item['文件大小'],
                'mtime': item['修改时间']
            })

        return tree

    def _render_html_tree(self, tree, title, root_path, use_relative_path=False, root_for_relative=None):
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
            background: #f5f7fa;
            padding: 30px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: #ffffff;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            padding: 30px 40px;
        }}
        .header {{
            border-bottom: 2px solid #e8ecf1;
            padding-bottom: 20px;
            margin-bottom: 25px;
        }}
        .header h1 {{ font-size: 26px; color: #1a2332; font-weight: 600; }}
        .header .subtitle {{ color: #7a8a9e; font-size: 14px; margin-top: 8px; }}
        .header .subtitle span {{
            background: #eef2f7;
            padding: 2px 12px;
            border-radius: 12px;
            margin-right: 10px;
        }}
        .tree {{ font-size: 15px; line-height: 2; }}
        .tree ul {{ list-style: none; padding-left: 24px; margin: 0; }}
        .tree li {{ position: relative; padding: 2px 0; }}
        .tree li::before {{
            content: '';
            position: absolute;
            left: -12px;
            top: 0;
            bottom: 0;
            border-left: 1.5px solid #d0d7e2;
        }}
        .tree li:last-child::before {{ height: 50%; }}
        .tree li .folder-toggle {{
            cursor: pointer;
            user-select: none;
            color: #1a2332;
            font-weight: 500;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 2px 8px;
            border-radius: 4px;
            transition: background 0.15s;
        }}
        .tree li .folder-toggle:hover {{ background: #f0f4ff; }}
        .tree li .folder-toggle .arrow {{
            display: inline-block;
            transition: transform 0.2s;
            font-size: 12px;
            color: #7a8a9e;
            width: 16px;
            text-align: center;
        }}
        .tree li .folder-toggle .arrow.open {{ transform: rotate(90deg); }}
        .tree li .folder-icon {{ color: #f5a623; }}
        .tree li .file-icon {{ color: #5b7a9a; margin-right: 6px; }}
        .tree li .file-item {{
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 2px 8px;
            border-radius: 4px;
            transition: background 0.15s;
            cursor: pointer;
        }}
        .tree li .file-item:hover {{ background: #f0f7ff; }}
        .tree li .file-item a {{
            color: #2b6eb3;
            text-decoration: none;
            font-weight: 400;
        }}
        .tree li .file-item a:hover {{ text-decoration: underline; }}
        .tree li .file-meta {{
            color: #9aa9bb;
            font-size: 12px;
            margin-left: 12px;
            font-weight: 400;
        }}
        .tree .collapsed > ul {{ display: none; }}
        .footer {{
            margin-top: 30px;
            padding-top: 18px;
            border-top: 1px solid #e8ecf1;
            color: #9aa9bb;
            font-size: 13px;
            text-align: center;
        }}
        .search-box {{
            margin-bottom: 20px;
            display: flex;
            gap: 10px;
        }}
        .search-box input {{
            flex: 1;
            padding: 10px 16px;
            border: 1.5px solid #d0d7e2;
            border-radius: 8px;
            font-size: 14px;
            outline: none;
            transition: border 0.2s;
        }}
        .search-box input:focus {{ border-color: #2b6eb3; }}
        .search-box .stats {{
            color: #7a8a9e;
            font-size: 14px;
            display: flex;
            align-items: center;
            white-space: nowrap;
        }}
        @media (max-width: 640px) {{
            body {{ padding: 12px; }}
            .container {{ padding: 16px; }}
            .tree ul {{ padding-left: 16px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📁 {title}</h1>
            <div class="subtitle">
                <span>📂 根目录: {root_path}</span>
                <span>📄 共 {len(self.file_data)} 个文件</span>
                <span>🕐 生成时间: {now}</span>
            </div>
        </div>
        <div class="search-box">
            <input type="text" id="searchInput" placeholder="🔍 输入文件名搜索..." oninput="filterTree()">
            <span class="stats" id="statsDisplay">显示全部</span>
        </div>
        <div class="tree" id="treeContainer">
'''
        html += self._render_tree_nodes(
            tree, 0,
            use_relative_path=use_relative_path,
            root_for_relative=root_for_relative
        )

        html += '''
        </div>
        <div class="footer">档案智盘 v1.0 · 点击文件名可打开或下载文件</div>
    </div>
    <script>
        document.querySelectorAll('.folder-toggle').forEach(el => {
            el.addEventListener('click', function(e) {
                e.stopPropagation();
                const li = this.closest('li');
                li.classList.toggle('collapsed');
                const arrow = this.querySelector('.arrow');
                if (arrow) arrow.classList.toggle('open');
            });
        });
        function filterTree() {
            const keyword = document.getElementById('searchInput').value.trim().toLowerCase();
            const stats = document.getElementById('statsDisplay');
            const tree = document.getElementById('treeContainer');
            if (!keyword) {
                tree.querySelectorAll('li').forEach(li => li.style.display = '');
                tree.querySelectorAll('li.collapsed').forEach(li => {
                    li.classList.remove('collapsed');
                    const arrow = li.querySelector('.arrow');
                    if (arrow) arrow.classList.add('open');
                });
                stats.textContent = '显示全部';
                return;
            }
            let matchCount = 0;
            const fileItems = tree.querySelectorAll('.file-item');
            fileItems.forEach(item => {
                const text = item.textContent.toLowerCase();
                const isMatch = text.includes(keyword);
                const li = item.closest('li');
                if (isMatch) {
                    li.style.display = '';
                    let parent = li.parentElement;
                    while (parent && parent.closest) {
                        const parentLi = parent.closest('li');
                        if (parentLi) {
                            parentLi.style.display = '';
                            parentLi.classList.remove('collapsed');
                            const arrow = parentLi.querySelector('.arrow');
                            if (arrow) arrow.classList.add('open');
                        }
                        parent = parent.parentElement;
                    }
                    matchCount++;
                }
            });
            const allLis = tree.querySelectorAll('li');
            allLis.forEach(li => {
                const fileItemsInLi = li.querySelectorAll('.file-item');
                if (fileItemsInLi.length > 0) {
                    let hasVisible = false;
                    fileItemsInLi.forEach(fi => {
                        if (fi.closest('li').style.display !== 'none') hasVisible = true;
                    });
                    if (!hasVisible) {
                        const allFileItems = li.querySelectorAll('.file-item');
                        let anyVisible = false;
                        allFileItems.forEach(fi => {
                            if (fi.closest('li').style.display !== 'none') anyVisible = true;
                        });
                        if (!anyVisible) li.style.display = 'none';
                    }
                }
            });
            stats.textContent = `找到 ${matchCount} 个匹配文件`;
        }
    </script>
</body>
</html>
'''
        return html

    def _render_tree_nodes(self, node, depth, use_relative_path=False, root_for_relative=None):
        html = ''
        indent = '  ' * depth

        for key, value in node.items():
            if key == '_files':
                continue

            html += f'\n{indent}<li>'
            html += f'''
            <div class="folder-toggle">
                <span class="arrow open">▶</span>
                <span class="folder-icon">📁</span>
                {key}
            </div>
            <ul>
'''
            if isinstance(value, dict):
                html += self._render_tree_nodes(
                    value, depth + 1,
                    use_relative_path=use_relative_path,
                    root_for_relative=root_for_relative
                )

            if '_files' in value and value['_files']:
                for file_info in value['_files']:
                    file_name = file_info['name']
                    file_path = file_info['path']
                    file_size = file_info.get('size', '')
                    file_mtime = file_info.get('mtime', '')

                    if use_relative_path and root_for_relative:
                        try:
                            rel = os.path.relpath(file_path, root_for_relative)
                            href = f"./文件/{rel.replace(os.sep, '/')}"
                        except ValueError:
                            href = f"file:///{file_path.replace(os.sep, '/')}"
                    else:
                        href = f"file:///{file_path.replace(os.sep, '/')}"

                    html += f'''
                <li>
                    <span class="file-item">
                        <span class="file-icon">📄</span>
                        <a href="{href}" target="_blank">{file_name}</a>
                        <span class="file-meta">{file_size} · {file_mtime}</span>
                    </span>
                </li>
'''

            html += f'\n{indent}</ul>\n{indent}</li>'

        return html

    # ==================== 打包发布 ====================
    def export_package(self, output_zip_path,
                       include_excel=True,
                       include_html=True,
                       include_files=True,
                       use_relative_path=True,
                       empty_folder_choice='skip'):
        if not self.file_data and not self.empty_folders:
            return False

        start_time = time.time()
        temp_dir = tempfile.mkdtemp(prefix="archive_pilot_")

        try:
            files_dir = os.path.join(temp_dir, "文件")

            if include_html:
                html_path = os.path.join(temp_dir, "文件清单.html")
                self.export_to_html(
                    html_path,
                    use_relative_path=use_relative_path,
                    root_for_relative=self.root_path
                )

            if include_excel:
                excel_path = os.path.join(temp_dir, "文件清单.xlsx")
                self.export_to_excel(
                    excel_path,
                    with_hyperlink=True,
                    include_empty_folders=(empty_folder_choice == 'include')
                )

            if include_files and self.file_data:
                for item in self.file_data:
                    src = item['完整路径']
                    rel_path = item['相对路径']
                    dst = os.path.join(files_dir, rel_path)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    try:
                        shutil.copy2(src, dst)
                    except (PermissionError, OSError) as e:
                        print(f"⚠️ 复制文件失败：{src} -> {e}")

            if empty_folder_choice == 'include' and self.empty_folders:
                for folder in self.empty_folders:
                    rel_path = os.path.relpath(folder, self.root_path)
                    if rel_path == '.':
                        continue
                    dst = os.path.join(files_dir, rel_path)
                    os.makedirs(dst, exist_ok=True)

            # 支持 UTF-8 中文文件名
            with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zi = zipfile.ZipInfo(arcname)
                        zi.flag_bits |= 0x800
                        with open(file_path, 'rb') as f:
                            data = f.read()
                        zf.writestr(zi, data)

            elapsed = time.time() - start_time
            print(f"✅ 打包完成！耗时：{elapsed:.2f} 秒")
            return True

        except Exception as e:
            print(f"❌ 打包失败：{e}")
            return False

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ==================== ⭐ 新增：仅导出文件夹结构（修复路径匹配） ====================
    def export_folders_only(self, output_path, root_dir, include_hidden=False, include_paths=None):
        """
        仅导出文件夹结构（不含文件）到 Excel
        Args:
            output_path: 输出 Excel 文件路径
            root_dir: 要扫描的根目录
            include_hidden: 是否包含隐藏文件夹
            include_paths: 要导出的文件夹路径列表（None 表示全部）
        Returns:
            bool: 是否成功
        """
        try:
            root_path = Path(root_dir).resolve()
            if not root_path.is_dir():
                return False

            # ⭐ 统一路径格式：标准化 include_paths 中的路径
            if include_paths is not None:
                include_paths = [os.path.normpath(p) for p in include_paths]

            folder_list = []
            for item in root_path.rglob('*'):
                if item.is_dir():
                    if not include_hidden and item.name.startswith('.'):
                        continue
                    # 如果指定了 include_paths，只导出指定的文件夹
                    if include_paths is not None:
                        norm_path = os.path.normpath(str(item))
                        if norm_path not in include_paths:
                            continue
                    rel_path = str(item.relative_to(root_path))
                    try:
                        mtime = datetime.fromtimestamp(item.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        mtime = ''
                    folder_list.append({
                        '文件夹路径': rel_path if rel_path != '.' else '',
                        '完整路径': str(item),
                        '文件夹名': item.name,
                        '修改时间': mtime,
                        '层级深度': len(rel_path.split(os.sep)) if rel_path != '.' else 0
                    })

            if not folder_list:
                return False

            folder_list.sort(key=lambda x: (x['层级深度'], x['文件夹名']))

            df = pd.DataFrame(folder_list)
            df.to_excel(output_path, index=False)
            return True
        except Exception as e:
            print(f"导出文件夹结构失败：{e}")
            return False